import tkinter as tk
import os
from typing import Dict
from openpyxl.cell.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import random

class TkinterApp:
    buttons: Dict[str, tk.Button] = {}
    labels: Dict[str, tk.Label] = {}

    def __init__(self, appname: str = "TkinterApp", width: int = 400, heigth: int = 400):
        self.root = tk.Tk()
        self.appname = appname
        self.root.title(appname)
        if os.name == "nt":
            self.root.wm_iconbitmap(bitmap="python3.ico")
        else:
            self.root.wm_iconbitmap(bitmap="@python3.xbm")
        self.root.iconphoto(self.root._w, tk.PhotoImage(file='python3.png'))
        self.root.geometry(f"{width}x{heigth}")
        
        self.create_buttons_labels('A')
        self.create_buttons_labels('B')
        
        self.save_button = tk.Button(self.root, text="Save", command=self.save_wb)
        self.save_button.pack(pady=10)

        self.append_button = tk.Button(self.root, text="Append random values", command=self.append_random)
        self.append_button.pack(pady=10)

        self.append_label = tk.Label(self.root, text="")
        self.append_label.pack(pady=10)

    def append_random(self):
        letter = random.choice('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
        number = random.random()
        self.append_label.config(text=f"{letter} | {number}")
        self.append_values([letter, number])

    def create_buttons_labels(self, column: str):
        self.buttons[column] = tk.Button(
            self.root, text=f"Get column {column}", command=lambda: self.get_column(column))
        self.buttons[column].pack(pady=10)
        self.labels[column] = tk.Label(self.root, text="")
        self.labels[column].pack(pady=10)

    def get_cell_value(self, cell: Cell) -> str:
        return str(cell.value)

    def get_column(self, column: str):
        cells_values = list(map(self.get_cell_value, self.ws[column]))
        self.labels[column].config(text="\n".join(cells_values))
    
    def append_values(self, values: list):
        self.ws.append(values)

    def load_excel(self, filename: str, read_only=False):
        self.filename = filename
        self.wb = Workbook()
        self.wb = load_workbook(filename, read_only)
        self.ws = self.wb.active

    def save_wb(self):
        self.wb.save(self.filename)

    def start(self):
        self.root.mainloop()


def main():
    app = TkinterApp("ExcelApp", 500, 800)
    app.load_excel('sample.xlsx')
    app.start()


if __name__ == "__main__":
    main()
